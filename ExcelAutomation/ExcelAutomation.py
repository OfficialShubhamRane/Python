import openpyxl as xl


def prcess_workbook(fileName, sheetName):
    wb = xl.load_workbook(fileName);
    sheet = wb[sheetName];

    for i in range(2, sheet.max_row + 1):
        print(sheet.cell(i, 3).value * 0.9);
        sheet.cell(i, 4).value = sheet.cell(i, 3).value * 0.9;

    sheet.cell(1, 4).value = "Discount(10%)";

    # Overriding current spreadsheet with changes
    wb.save(fileName)

prcess_workbook("transactions.xlsx","Sheet1");
